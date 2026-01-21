"""Integration tests for Excel import functionality."""

import sys
import os
import tempfile
from pathlib import Path

# Add src to path
sys.path.insert(0, str(Path(__file__).parent.parent / 'src'))


def create_test_excel_valid(file_path: Path):
    """Create a test Excel file with valid data."""
    try:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        # Headers
        headers = [
            "First Name", "Last Name", "Email", "Phone",
            "External ID", "Status", "Workspace", "Role",
            "Contract", "Entry Date"
        ]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # Test data (3 rows)
        test_data = [
            ["Jean", "Dupont", "jean.dupont@example.com", "06 12 34 56 78", "WMS-001", "Actif", "Zone A", "Cariste", "CDI", "15/01/2025"],
            ["Marie", "Martin", "marie.martin@example.com", "06 23 45 67 89", "WMS-002", "Actif", "Zone B", "Magasinier", "CDD", "16/01/2025"],
            ["Pierre", "Bernard", "pierre.bernard@example.com", "06 34 56 78 90", "WMS-003", "Inactif", "Zone C", "PrEparateur", "Interim", "17/01/2025"],
        ]

        for row_idx, row_data in enumerate(test_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(file_path)
        wb.close()  # Explicitly close before returning
        return True
    except Exception as e:
        print(f"  [FAIL] Could not create test Excel: {e}")
        return False


def test_full_import_flow():
    """Test 1: Complete import flow with valid data."""
    print("[TEST 1] Testing complete import flow...")

    try:
        from excel_import import ExcelImporter
        from database.connection import database
        from employee.models import Employee

        # Create temporary Excel file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            # Create test Excel
            if not create_test_excel_valid(tmp_path):
                return False

            # Create importer
            importer = ExcelImporter(tmp_path)

            # Validate file
            is_valid, error_msg = importer.validate_file()
            if not is_valid:
                print(f"  [FAIL] File validation failed: {error_msg}")
                return False
            print("  [OK] File validation passed")

            # Parse file
            rows = importer.parse_file()
            if len(rows) != 3:
                print(f"  [FAIL] Expected 3 rows, got {len(rows)}")
                return False
            print(f"  [OK] Parsed {len(rows)} rows")

            # Preview data
            preview = importer.preview(max_rows=3)
            if preview['total_rows'] != 3:
                print(f"  [FAIL] Preview shows wrong row count")
                return False
            print("  [OK] Preview generated correctly")

            # Note: We don't actually import to avoid modifying database
            # In a real test, you would use a test database
            print("  [OK] Import flow validated (database import skipped)")

            # Close importer to release file lock
            importer.close()

            return True

        finally:
            # Clean up temp file
            if tmp_path.exists():
                tmp_path.unlink()

    except Exception as e:
        print(f"  [FAIL] Test failed: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_template_generation():
    """Test 2: Template generation."""
    print("\n[TEST 2] Testing template generation...")

    try:
        from excel_import import ExcelTemplateGenerator

        # Create temp file for template
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            # Generate template
            generator = ExcelTemplateGenerator()
            generator.generate_template(tmp_path)

            if not tmp_path.exists():
                print("  [FAIL] Template file not created")
                return False
            print(f"  [OK] Template file created")

            # Verify file can be opened
            from openpyxl import load_workbook
            wb = load_workbook(tmp_path)

            # Check sheets
            if "Instructions" not in wb.sheetnames:
                wb.close()
                print("  [FAIL] Instructions sheet missing")
                return False
            print("  [OK] Instructions sheet present")

            if "Data" not in wb.sheetnames:
                wb.close()
                print("  [FAIL] Data sheet missing")
                return False
            print("  [OK] Data sheet present")

            # Check headers in Data sheet
            ws = wb["Data"]
            headers = [ws.cell(1, col).value for col in range(1, 11)]

            # Template adds " *" to required columns (except Entry Date which already has *)
            expected_headers = [
                "First Name *", "Last Name *", "Email", "Phone",
                "External ID", "Status *", "Workspace *", "Role *",
                "Contract *", "Entry Date"
            ]

            if headers != expected_headers:
                wb.close()
                print(f"  [FAIL] Headers mismatch: {headers}")
                print(f"  Expected: {expected_headers}")
                return False
            print("  [OK] All headers correct (with required markers)")

            wb.close()
            return True

        finally:
            # Clean up
            if tmp_path.exists():
                tmp_path.unlink()

    except Exception as e:
        print(f"  [FAIL] Test failed: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_sample_file_generation():
    """Test 3: Sample file generation."""
    print("\n[TEST 3] Testing sample file generation...")

    try:
        from excel_import import ExcelTemplateGenerator

        # Create temp file for sample
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            # Generate sample with 5 employees
            generator = ExcelTemplateGenerator()
            generator.generate_sample_file(tmp_path, num_employees=5)

            if not tmp_path.exists():
                print("  [FAIL] Sample file not created")
                return False
            print("  [OK] Sample file created")

            # Verify file can be opened
            from openpyxl import load_workbook
            wb = load_workbook(tmp_path)
            ws = wb.active

            # Check row count (header + 5 data rows = 6)
            if ws.max_row != 6:
                print(f"  [FAIL] Expected 6 rows, got {ws.max_row}")
                return False
            print("  [OK] Correct number of rows (header + 5 data rows)")

            # Verify data in first data row
            first_name = ws.cell(2, 1).value
            if not first_name:
                print("  [FAIL] No data in first row")
                return False
            print(f"  [OK] Sample data present (e.g., {first_name})")

            wb.close()
            return True

        finally:
            # Clean up
            if tmp_path.exists():
                tmp_path.unlink()

    except Exception as e:
        print(f"  [FAIL] Test failed: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_validation_error_handling():
    """Test 4: Validation error handling."""
    print("\n[TEST 4] Testing validation error handling...")

    try:
        from excel_import import ExcelImporter
        from openpyxl import Workbook

        # Create temp file with missing required column
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            wb = Workbook()
            ws = wb.active

            # Headers - missing "Entry Date" (required)
            headers = [
                "First Name", "Last Name", "Email", "Phone",
                "External ID", "Status", "Workspace", "Role",
                "Contract"
                # Entry Date is missing!
            ]
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)

            # Add one data row
            ws.cell(row=2, column=1, value="Jean")
            ws.cell(row=2, column=2, value="Dupont")

            wb.save(tmp_path)
            wb.close()  # Close before trying to import

            # Try to import
            importer = ExcelImporter(tmp_path)
            is_valid, error_msg = importer.validate_file()

            # Close importer to release file lock
            importer.close()

            if is_valid:
                print("  [FAIL] Validation should have failed for missing column")
                return False

            if "Entry Date" not in error_msg and "required" not in error_msg.lower():
                print(f"  [FAIL] Error message should mention missing column: {error_msg}")
                return False

            print(f"  [OK] Validation correctly failed: {error_msg}")
            return True

        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    except Exception as e:
        print(f"  [FAIL] Test failed: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_date_format_handling():
    """Test 5: Date format handling."""
    print("\n[TEST 5] Testing date format handling...")

    try:
        from excel_import import ExcelImporter
        from datetime import date

        # Test French format
        result = ExcelImporter._parse_date("15/01/2025")
        if result != date(2025, 1, 15):
            print(f"  [FAIL] French format failed: {result}")
            return False
        print("  [OK] DD/MM/YYYY format parsed correctly")

        # Test ISO format
        result = ExcelImporter._parse_date("2025-01-15")
        if result != date(2025, 1, 15):
            print(f"  [FAIL] ISO format failed: {result}")
            return False
        print("  [OK] YYYY-MM-DD format parsed correctly")

        # Test invalid format
        result = ExcelImporter._parse_date("15-01-2025")  # Wrong separator
        if result is not None:
            print(f"  [FAIL] Invalid format should return None, got {result}")
            return False
        print("  [OK] Invalid format returns None")

        # Test empty string
        result = ExcelImporter._parse_date("")
        if result is not None:
            print(f"  [FAIL] Empty string should return None, got {result}")
            return False
        print("  [OK] Empty string returns None")

        return True

    except Exception as e:
        print(f"  [FAIL] Test failed: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_string_cleaning():
    """Test 6: String cleaning functionality."""
    print("\n[TEST 6] Testing string cleaning...")

    try:
        from excel_import import ExcelImporter

        # Test basic trimming and capitalization
        result = ExcelImporter._clean_string("  jean  ")
        if result != "Jean":
            print(f"  [FAIL] Expected 'Jean', got '{result}'")
            return False
        print("  [OK] Trims and capitalizes correctly")

        # Test None
        result = ExcelImporter._clean_string(None)
        if result is not None:
            print(f"  [FAIL] None should return None, got {result}")
            return False
        print("  [OK] None returns None")

        # Test empty string
        result = ExcelImporter._clean_string("")
        if result is not None:
            print(f"  [FAIL] Empty string should return None, got {result}")
            return False
        print("  [OK] Empty string returns None")

        # Test single character
        result = ExcelImporter._clean_string("a")
        if result != "A":
            print(f"  [FAIL] Single char 'a' should become 'A', got '{result}'")
            return False
        print("  [OK] Single character capitalized")

        # Test already capitalized
        result = ExcelImporter._clean_string("Jean")
        if result != "Jean":
            print(f"  [FAIL] Already capitalized should stay same, got '{result}'")
            return False
        print("  [OK] Already capitalized stays same")

        return True

    except Exception as e:
        print(f"  [FAIL] Test failed: {e}")
        import traceback
        traceback.print_exc()
        return False


def run_all_tests():
    """Run all integration tests."""
    print("=" * 60)
    print("EXCEL IMPORT INTEGRATION TESTS")
    print("=" * 60)
    print()

    tests = [
        test_full_import_flow,
        test_template_generation,
        test_sample_file_generation,
        test_validation_error_handling,
        test_date_format_handling,
        test_string_cleaning,
    ]

    results = []
    for test in tests:
        try:
            result = test()
            results.append(result)
        except Exception as e:
            print(f"\n[ERROR] Test crashed: {e}")
            import traceback
            traceback.print_exc()
            results.append(False)

    # Summary
    print("\n" + "=" * 60)
    print("TEST SUMMARY")
    print("=" * 60)
    passed = sum(results)
    total = len(results)

    print(f"Tests run: {total}")
    print(f"Tests passed: {passed}")
    print(f"Tests failed: {total - passed}")
    print(f"Success rate: {(passed/total)*100:.1f}%")

    if passed == total:
        print("\n[OK] ALL TESTS PASSED")
        return 0
    else:
        print(f"\n[FAIL] {total - passed} TEST(S) FAILED")
        return 1


if __name__ == "__main__":
    exit_code = run_all_tests()
    sys.exit(exit_code)
